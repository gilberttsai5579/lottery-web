"""
Excel export functionality for lottery results
"""
import os
from datetime import datetime
from typing import List, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..models import LotteryResult, Comment


class ExcelExporter:
    """
    Excel exporter for lottery results
    """
    
    def __init__(self, output_dir: str = "output"):
        """
        Initialize Excel exporter
        
        Args:
            output_dir: Directory to save Excel files
        """
        self.output_dir = output_dir
        self.logger = self._setup_logger()
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl not available. Excel export will be limited.")
    
    def _setup_logger(self) -> logging.Logger:
        """Setup logger for Excel exporter"""
        logger = logging.getLogger("ExcelExporter")
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        return logger
    
    def export_lottery_result(self, result: LotteryResult, filename: Optional[str] = None) -> str:
        """
        Export lottery result to Excel file
        
        Args:
            result: LotteryResult to export
            filename: Optional filename (auto-generated if not provided)
            
        Returns:
            Path to the exported file
            
        Raises:
            RuntimeError: If openpyxl is not available
        """
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV export
            return self._export_to_csv(result, filename)
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"lottery_result_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        self.logger.info(f"Exporting lottery result to {filepath}")
        
        # Create workbook
        wb = Workbook()
        
        # Create worksheets
        self._create_summary_sheet(wb, result)
        self._create_winners_sheet(wb, result)
        self._create_all_participants_sheet(wb, result)
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Save workbook
        wb.save(filepath)
        
        self.logger.info(f"Excel file exported successfully: {filepath}")
        return filepath
    
    def _create_summary_sheet(self, wb: Workbook, result: LotteryResult):
        """Create summary worksheet"""
        ws = wb.create_sheet("摘要資訊")
        
        # Headers
        headers = [
            ("抽獎資訊", ""),
            ("抽獎時間", result.timestamp.strftime("%Y-%m-%d %H:%M:%S")),
            ("貼文網址", result.post_url),
            ("平台", result.platform.upper()),
            ("抽獎模式", result.mode_name),
            ("", ""),
            ("參數設定", ""),
            ("中獎人數", str(result.winner_count)),
        ]
        
        if result.keyword:
            headers.append(("關鍵字", result.keyword))
        if result.mention_count_required > 1:
            headers.append(("需標註數量", str(result.mention_count_required)))
        
        headers.extend([
            ("", ""),
            ("統計資訊", ""),
            ("總留言數", str(result.total_comments)),
            ("總參與人數", str(result.total_participants)),
            ("符合條件人數", str(result.eligible_count)),
            ("實際中獎人數", str(len(result.winners))),
        ])
        
        # Write headers
        for i, (label, value) in enumerate(headers, 1):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
        
        # Style the worksheet
        self._style_summary_sheet(ws)
    
    def _create_winners_sheet(self, wb: Workbook, result: LotteryResult):
        """Create winners worksheet"""
        ws = wb.create_sheet("中獎名單")
        
        # Headers
        headers = ["排名", "用戶名", "留言內容", "標註數量", "按讚數", "頭像網址"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Winners data
        for i, winner in enumerate(result.winners, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=winner.username)
            ws.cell(row=row, column=3, value=winner.content)
            ws.cell(row=row, column=4, value=winner.mention_count())
            ws.cell(row=row, column=5, value=winner.likes_count)
            ws.cell(row=row, column=6, value=winner.avatar_url or "")
        
        # Style the worksheet
        self._style_data_sheet(ws, len(headers))
    
    def _create_all_participants_sheet(self, wb: Workbook, result: LotteryResult):
        """Create all participants worksheet"""
        ws = wb.create_sheet("所有參與者")
        
        # Headers
        headers = ["用戶名", "留言內容", "標註數量", "符合條件", "按讚數", "平台"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Participants data
        for i, participant in enumerate(result.all_participants, 1):
            row = i + 1
            is_eligible = participant in result.eligible_participants
            
            ws.cell(row=row, column=1, value=participant.username)
            ws.cell(row=row, column=2, value=participant.content)
            ws.cell(row=row, column=3, value=participant.mention_count())
            ws.cell(row=row, column=4, value="是" if is_eligible else "否")
            ws.cell(row=row, column=5, value=participant.likes_count)
            ws.cell(row=row, column=6, value=participant.platform.upper())
        
        # Style the worksheet
        self._style_data_sheet(ws, len(headers))
    
    def _style_summary_sheet(self, ws):
        """Apply styles to summary sheet"""
        # Header style
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Apply styles
        for row in ws.iter_rows():
            for cell in row:
                if cell.column == 1 and cell.value:  # Labels column
                    cell.font = header_font
                    if "資訊" in str(cell.value):
                        cell.fill = header_fill
                
                # Auto-adjust column width
                if cell.value:
                    column_letter = get_column_letter(cell.column)
                    current_width = ws.column_dimensions[column_letter].width or 10
                    new_width = len(str(cell.value)) + 2
                    ws.column_dimensions[column_letter].width = max(current_width, new_width)
    
    def _style_data_sheet(self, ws, num_columns: int):
        """Apply styles to data sheets"""
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styles
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Apply data styles and borders
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Auto-adjust column widths
        for col in range(1, num_columns + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # Set reasonable width limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _export_to_csv(self, result: LotteryResult, filename: Optional[str] = None) -> str:
        """
        Fallback CSV export when openpyxl is not available
        
        Args:
            result: LotteryResult to export
            filename: Optional filename
            
        Returns:
            Path to exported CSV file
        """
        import csv
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"lottery_result_{timestamp}.csv"
        
        filepath = os.path.join(self.output_dir, filename)
        
        self.logger.info(f"Exporting lottery result to CSV: {filepath}")
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # Summary information
            writer.writerow(["抽獎摘要"])
            writer.writerow(["抽獎時間", result.timestamp.strftime("%Y-%m-%d %H:%M:%S")])
            writer.writerow(["貼文網址", result.post_url])
            writer.writerow(["平台", result.platform.upper()])
            writer.writerow(["抽獎模式", result.mode_name])
            writer.writerow(["中獎人數", result.winner_count])
            
            if result.keyword:
                writer.writerow(["關鍵字", result.keyword])
            if result.mention_count_required > 1:
                writer.writerow(["需標註數量", result.mention_count_required])
            
            writer.writerow(["總留言數", result.total_comments])
            writer.writerow(["總參與人數", result.total_participants])
            writer.writerow(["符合條件人數", result.eligible_count])
            writer.writerow(["實際中獎人數", len(result.winners)])
            writer.writerow([])
            
            # Winners
            writer.writerow(["中獎名單"])
            writer.writerow(["排名", "用戶名", "留言內容", "標註數量", "按讚數"])
            
            for i, winner in enumerate(result.winners, 1):
                writer.writerow([
                    i,
                    winner.username,
                    winner.content,
                    winner.mention_count(),
                    winner.likes_count
                ])
        
        self.logger.info(f"CSV file exported successfully: {filepath}")
        return filepath
    
    def get_file_path(self, filename: str) -> str:
        """
        Get full file path for a filename
        
        Args:
            filename: Filename
            
        Returns:
            Full file path
        """
        return os.path.join(self.output_dir, filename)
    
    def file_exists(self, filename: str) -> bool:
        """
        Check if exported file exists
        
        Args:
            filename: Filename to check
            
        Returns:
            True if file exists
        """
        filepath = self.get_file_path(filename)
        return os.path.exists(filepath)